import os
import pandas as pd
import zipfile
import rarfile
import py7zr
import tempfile
import shutil
import pdfplumber
from docx import Document
from openpyxl import load_workbook

def extract_text_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()

def extract_text_from_docx(file_path):
    doc = Document(file_path)
    return '\n'.join([para.text for para in doc.paragraphs])

def extract_text_from_xlsx(file_path):
    wb = load_workbook(file_path, data_only=True)
    text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = ' '.join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                text.append(row_text)
    return '\n'.join(text)

def extract_text_from_pdf(file_path):
    text = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)
    return '\n'.join(text)

def handle_archive(archive_path, temp_extract_dir):
    extracted_files = []
    try:
        if archive_path.endswith('.zip'):
            with zipfile.ZipFile(archive_path, 'r') as zip_ref:
                zip_ref.extractall(temp_extract_dir)
        elif archive_path.endswith('.rar'):
            with rarfile.RarFile(archive_path, 'r') as rar_ref:
                rar_ref.extractall(temp_extract_dir)
        elif archive_path.endswith('.7z'):
            with py7zr.SevenZipFile(archive_path, 'r') as sz_ref:
                sz_ref.extractall(temp_extract_dir)
        else:
            return []
        for root, dirs, files in os.walk(temp_extract_dir):
            for file in files:
                extracted_files.append(os.path.join(root, file))
    except Exception as e:
        print(f"Ошибка при распаковке {archive_path}: {e}")
    return extracted_files

def crawl_and_extract(root_dir):
    results = []
    processed = set()

    for root, dirs, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(root, file)
            if file_path in processed:
                continue

            rel_path = os.path.relpath(file_path, root_dir)
            file_ext = os.path.splitext(file)[1].lower()
            content = None

            try:
                if file_ext == '.txt':
                    content = extract_text_from_txt(file_path)
                elif file_ext == '.docx':
                    content = extract_text_from_docx(file_path)
                elif file_ext == '.xlsx':
                    content = extract_text_from_xlsx(file_path)
                elif file_ext == '.pdf':
                    content = extract_text_from_pdf(file_path)
                elif file_ext in ['.zip', '.rar', '.7z']:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        extracted = handle_archive(file_path, tmpdir)
                        for ext_file in extracted:
                            if ext_file not in processed:
                                ext_rel = os.path.relpath(ext_file, tmpdir)
                                ext_ext = os.path.splitext(ext_file)[1].lower()
                                ext_content = None
                                if ext_ext == '.txt':
                                    ext_content = extract_text_from_txt(ext_file)
                                elif ext_ext == '.docx':
                                    ext_content = extract_text_from_docx(ext_file)
                                elif ext_ext == '.xlsx':
                                    ext_content = extract_text_from_xlsx(ext_file)
                                elif ext_ext == '.pdf':
                                    ext_content = extract_text_from_pdf(ext_file)
                                else:
                                    ext_content = "[Не поддерживаемый формат]"
                                
                                results.append({
                                    'file_path': f"{rel_path} -> {ext_rel}",
                                    'file_name': os.path.basename(ext_file),
                                    'file_type': ext_ext,
                                    'content': ext_content
                                })
                                processed.add(ext_file)
                    processed.add(file_path)
                    continue
                else:
                    content = "[Не поддерживаемый формат]"
            except Exception as e:
                content = f"[Ошибка парсинга: {e}]"

            if file_ext not in ['.zip', '.rar', '.7z']:
                results.append({
                    'file_path': rel_path,
                    'file_name': file,
                    'file_type': file_ext,
                    'content': content
                })
                processed.add(file_path)

    return results

if __name__ == "__main__":
    data = crawl_and_extract('test_storage')
    df = pd.DataFrame(data)
    os.makedirs('data', exist_ok=True)
    df.to_csv('data/crawled_data.csv', index=False, encoding='utf-8-sig')
    print(f"✅ Данные сохранены в data/crawled_data.csv (всего {len(df)} записей)")